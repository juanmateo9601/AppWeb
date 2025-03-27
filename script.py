import streamlit as st
import pandas as pd
import io
from streamlit_pdf_viewer import pdf_viewer
import re
import json
from io import StringIO, BytesIO
import base64
from openpyxl import load_workbook
import os
import warnings
import streamlit.components.v1 as components
from PIL import Image
from datetime import datetime
from openpyxl.styles import Alignment
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

ruta_plantilla = "Plantilla_Turbo_Final.xlsx"

TECNICOS = [
    {
        "PROFESIONAL": "Lazaro Alpidio Londoño Londoño",
        "CARGO": "Profesional Tecnico",
        "CEDULA": "70.195.935"
    },
    {
        "PROFESIONAL": "Vannesa Orozco Perez",
        "CARGO": "Profesional Tecnico",
        "CEDULA": "1.020.438.209"
    },
    {
        "PROFESIONAL": "Binis Shirley Viana Padilla",
        "CARGO": "Profesional Tecnico",
        "CEDULA": "43.655.326"
    },
    {
        "PROFESIONAL": "Julian Mauricio Madrid",
        "CARGO": "Profesional Tecnico",
        "CEDULA": "1.035.865.095"
    },
    {
        "PROFESIONAL": "Paola Andrea Álvarez Ramírez",
        "CARGO": "Profesional Tecnico",
        "CEDULA": "44.007.552"
    },
    {
        "PROFESIONAL": "Felipe Agudelo Espitia",
        "CARGO": "Profesional Tecnico",
        "CEDULA": "1.020.411.138"
    },
    {
        "PROFESIONAL": "Jose Rafael Oliveros Mora",
        "CARGO": "Profesional Tecnico",
        "CEDULA": "1.082.947.743"
    },
    {
        "PROFESIONAL": "Santiago Zapata Zuluaga",
        "CARGO": "Profesional Tecnico",
        "CEDULA": "1.037.662.712"
    }
]

def obtener_tabla_habitaciones():
    if "costos_excel" in st.session_state:
        df_costos = st.session_state["costos_excel"].copy()

        columnas_exportar = [
            "Item",
            "ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS",
            "Unidad",
            "Valor Unitario ofertado (**)"
        ]
        df_intermedio = df_costos[columnas_exportar].copy()

        categoria_actual = None
        categorias = []
        
        for _, row in df_intermedio.iterrows():
            actividad = str(row["ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS"])
            if actividad.isupper():
                categoria_actual = actividad
                categorias.append(categoria_actual)
            else:
                categorias.append(categoria_actual)
        
        df_intermedio["Categoria"] = categorias

        if "resultados_csv" in st.session_state:
            habitaciones_procesadas = [
                habitacion
                for habitacion in st.session_state["resultados_csv"].keys()
                if "piso" not in habitacion.lower()
            ]

            for habitacion in habitaciones_procesadas:
                df_intermedio[habitacion] = 0.0
                for i, row in df_intermedio.iterrows():
                    actividad = row["ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS"]
                    cantidad_key = f"cantidad_{habitacion}_{actividad}"
                    if cantidad_key in st.session_state:
                        val = st.session_state[cantidad_key]
                        if val == "" or val is None:
                            val_float = 0.0
                        else:
                            try:
                                val_float = float(val)
                            except (ValueError, TypeError):
                                val_float = 0.0
                        df_intermedio.at[i, habitacion] = val_float

            # ⭕️ Esta línea faltaba y causaba el primer error:
            df_intermedio["Total actividad"] = df_intermedio[habitaciones_procesadas].sum(axis=1)

            # 🟢 Corrección del nombre exacto de la columna:
            df_intermedio["Costo total"] = (
                df_intermedio["Total actividad"] *
                df_intermedio["Valor Unitario ofertado (**)"
            ])

            df_resumen = df_intermedio[[
                "Item",
                "Categoria",
                "ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS",
                "Unidad",
                "Valor Unitario ofertado (**)",
                "Total actividad",
                "Costo total"
            ]].copy()
            
            df_resumen = df_resumen.rename(columns={
                "Item": "N°",
                "ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS": "DESCRIPCIÓN",
                "Unidad": "UN",
                "Total actividad": "CANT INIC",
                "Valor Unitario ofertado (**)": "VR INIT",
                "Costo total": "VR TOTAL"
            })

        nueva_ruta = export_to_excel_pure(df_resumen, st.session_state.get("selected_tecnico"))
        st.session_state["export_excel"] = nueva_ruta


from openpyxl.styles import Font, Alignment  # Asegúrate de tener esto al inicio del archivo

def export_to_excel_pure(datos_resumen, selected_tecnico=None):
    """
    Llena la plantilla con las actividades > 0 desde un DataFrame (o lista de diccionarios),
    agrupándolas por la columna 'Categoria', inyecta los datos del beneficiario en el encabezado
    (según la cédula ingresada) y guarda automáticamente el archivo en la carpeta Downloads del usuario.
    
    Si se proporciona selected_tecnico, inyecta sus datos en las celdas B99, C100 y B101.
    """
    import os
    from pathlib import Path
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    # Comprobar si datos_resumen está vacío (para DataFrame)
    if datos_resumen is None or (hasattr(datos_resumen, "empty") and datos_resumen.empty):
        st.error("⚠️ datos_resumen está vacío. No hay datos para exportar.")
        return None

    st.write("🔍 Datos con Total actividad calculado:", datos_resumen)

    # Si datos_resumen es un DataFrame, lo convertimos a lista de diccionarios
    if hasattr(datos_resumen, "to_dict"):
        datos_filtrados = datos_resumen.to_dict(orient="records")
    else:
        datos_filtrados = datos_resumen

    # Filtrar actividades con "Total actividad" > 0
    datos_filtrados = [fila for fila in datos_filtrados if fila.get("CANT INIC", 0) > 0]
    if not datos_filtrados:
        st.warning("⚠️ No hay actividades con valor > 0. El Excel quedará vacío.")
        return None

    # Obtener la carpeta Downloads del usuario y definir el nombre del archivo
    downloads_folder = Path.home() / "Downloads"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nueva_ruta = downloads_folder / f"Reporte_Resultado_{timestamp}.xlsx"

    # Función para obtener la ruta de la plantilla
    def get_plantilla_path():
        plantilla_path = os.path.join(os.getcwd(), "Plantilla_Turbo_Final.xlsx")
        if not os.path.exists(plantilla_path):
            st.error(f"❌ No se encontró la plantilla: {plantilla_path}")
            return None
        return plantilla_path

    ruta_plantilla_local = get_plantilla_path()
    if ruta_plantilla_local is None:
        return None

    try:
        wb = load_workbook(ruta_plantilla_local)
    except Exception as e:
        st.error(f"Error al cargar la plantilla: {e}")
        return None

    ws = wb.active

    # Función para concatenar texto nuevo con el contenido original de la celda
    def append_to_cell(cell_ref, new_text):
        original = ws[cell_ref].value or ""
        if cell_ref == "E99":
            ws[cell_ref].value = f"{original}{new_text}"
        else:
            ws[cell_ref].value = f"{original}    {new_text}"

    # Inyectar datos del beneficiario (usando st.session_state)
    if "cedula_usuario" in st.session_state and "beneficiarios_excel" in st.session_state:
        cedula_dig = st.session_state["cedula_usuario"].strip()
        beneficiarios = st.session_state["beneficiarios_excel"]

        beneficiario_encontrado = None
        if hasattr(beneficiarios, "empty"):  # Si es un DataFrame
            df_benef = beneficiarios.copy()
            df_benef["C.C:"] = df_benef["C.C:"].astype(str)
            df_filtrado = df_benef[df_benef["C.C:"].str.strip() == cedula_dig]
            if not df_filtrado.empty:
                beneficiario_encontrado = df_filtrado.iloc[0].to_dict()
        else:
            beneficiario_encontrado = next((b for b in beneficiarios 
                                             if str(b.get("C.C:", "")).strip() == cedula_dig), None)

        if beneficiario_encontrado:
            append_to_cell("B7", beneficiario_encontrado.get("NOMBRE BENEFICIARIO:", ""))
            append_to_cell("E99", beneficiario_encontrado.get("NOMBRE BENEFICIARIO:", ""))
            append_to_cell("B8", beneficiario_encontrado.get("C.C:", ""))
            append_to_cell("E100", beneficiario_encontrado.get("C.C:", ""))
            append_to_cell("G7", beneficiario_encontrado.get("ID_HOGAR:", "N/A"))
            append_to_cell("B9", beneficiario_encontrado.get("TELEFONO:", ""))
            append_to_cell("D8", beneficiario_encontrado.get("DIRECCION:", ""))
        else:
            st.write(f"No se encontró la cédula {cedula_dig} en la base de beneficiarios.")

    # Agregar fecha sin sobrescribir la celda
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    append_to_cell("G4", fecha_actual)
    append_to_cell("D9", fecha_actual)

    # Inyectar datos del técnico seleccionado, si se proporcionó
    if selected_tecnico:
        ws["B99"] = selected_tecnico.get("PROFESIONAL", "")
        ws["C100"] = selected_tecnico.get("CEDULA", "")
        ws["B101"] = selected_tecnico.get("CARGO", "")

    # Identificar celdas combinadas para evitar escribir en ellas
    celdas_combinadas = set()
    for merged_range in ws.merged_cells.ranges:
        for row in ws[merged_range.coord]:
            for cell in row:
                celdas_combinadas.add(cell.coordinate)

    # Obtener categorías únicas a partir de los datos filtrados
    categorias_unicas = list({fila["Categoria"] for fila in datos_filtrados if fila.get("Categoria")})
    current_row = 14

    for cat in categorias_unicas:
        ws[f"B{current_row}"] = cat
        ws[f"B{current_row}"].alignment = Alignment(horizontal="left", indent=2)
        ws[f"B{current_row}"].fill = PatternFill("solid", fgColor="D3D3D3")
        ws[f"B{current_row}"].font = Font(color="000000", bold=True)
        current_row += 1

        # Filtrar actividades de la categoría actual
        actividades = [f for f in datos_filtrados if f.get("Categoria") == cat]
        # Ajustar columnas: N° en B, DESCRIPCIÓN en C, UN en D, CANT INIC en E, VR INIT en F, VR TOTAL en G
        col_map = ["B", "C", "D", "E", "F", "G"]

        for act in actividades:
            valores = [
                act.get("N°", ""),
                act.get("DESCRIPCIÓN", ""),
                act.get("UN", ""),
                act.get("CANT INIC", 0.0),
                act.get("VR INIT (**)", 0.0),
                act.get("VR TOTAL", 0.0)
            ]
            for col, val in zip(col_map, valores):
                celda = f"{col}{current_row}"
                if celda not in celdas_combinadas:
                    if col in ["F", "G"]:
                        try:
                            ws[celda].value = float(val)
                            ws[celda].number_format = '"$"#,##0'
                        except (ValueError, TypeError):
                            ws[celda].value = 0.0
                    else:
                        ws[celda].value = val
                        if col == "B":
                            ws[celda].font = Font(color="000000", bold=False)
                            ws[celda].hyperlink = None
            current_row += 1

        current_row += 1

    ws["G77"] = "=SUM(G15:G76)"
    ws["G77"].number_format = '"$"#,##0'
    
    ws["G81"] = "=G77*0.12"
    ws["G81"].number_format = '"$"#,##0.00'
    
    ws["G82"] = "=G77*0.016"
    ws["G82"].number_format = '"$"#,##0.00'
    
    ws["G83"] = "=G77+G81+G82"
    ws["G83"].number_format = '"$"#,##0.00'
    
    # Replicar el valor de G83 en G85 considerando celdas fusionadas
    target = "G85"
    found = False
    for merged_range in ws.merged_cells.ranges:
        if target in merged_range:
            ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = "=G83"
            ws.cell(row=merged_range.min_row, column=merged_range.min_col).number_format = '"$"#,##0.00'
            found = True
            break
    if not found:
        ws[target].value = "=G83"
        ws[target].number_format = '"$"#,##0.00'

    try:
        wb.save(nueva_ruta)
        st.write(f"✅ Reporte guardado automáticamente en: {nueva_ruta}")
    except Exception as e:
        st.error(f"❌ Error al guardar el archivo Excel: {e}")
        return None

    return str(nueva_ruta)




def procesar_csv_bytes(file_bytes: BytesIO):
    """
    Procesa un archivo CSV desde un BytesIO y devuelve un diccionario con las tablas encontradas.

    Args:
        file_bytes (BytesIO): Archivo CSV en memoria.

    Returns:
        tuple: Un diccionario con las tablas y un código de estado HTTP.
    """
    try:
        content = file_bytes.getvalue().decode('utf-8', errors='replace')

        raw_sections = re.split(r'\n\s*\n+', content)
        sections = [sec.strip() for sec in raw_sections if sec.strip()]
        
        tablas = {}
        for idx, section in enumerate(sections, start=1):
            lines = section.split('\n')

            if len(lines) == 1:
                tablas[f"tabla_{idx}"] = {"titulo": lines[0]}
                continue
            
            if all(':' in line for line in lines if line.strip()):
                data = {key.strip(): value.strip().strip(',')
                        for line in lines if (parts := line.split(':', 1)) and len(parts) == 2
                        for key, value in [parts]}
                tablas[f"tabla_{idx}"] = data
                continue
            
            try:
                read_csv_kwargs = {"encoding": "utf-8"}
                if pd.__version__ >= "1.3.0":
                    read_csv_kwargs["on_bad_lines"] = "skip"
                else:
                    read_csv_kwargs["error_bad_lines"] = False
                
                df = pd.read_csv(StringIO(section), **read_csv_kwargs)
                
                if not df.empty:
                    df.columns = df.columns.str.strip()
                    tablas[f"tabla_{idx}"] = df
                    continue
            except pd.errors.ParserError:
                pass  

            data = {f"columna_{i}": [part.strip() for part in line.split(',')] 
                    if ',' in line else line.strip() for i, line in enumerate(lines)}
            tablas[f"tabla_{idx}"] = data

        return tablas, 200
    except UnicodeDecodeError:
        return {"error": "Error al leer el archivo, posible problema de codificación"}, 400
    except Exception as e:
        return {"error": f"Error al procesar el archivo CSV: {str(e)}"}, 500

def calcular_propiedades_habitacion(tablas):
    """
    Calcula valores para cada habitación en las tablas encontradas.

    Args:
        tablas (dict): Diccionario de tablas procesadas.

    Returns:
        dict: JSON con los resultados en formato de diccionario.
    """
    resultados = {}

    for tabla_key, value in tablas.items():
        if isinstance(value, pd.DataFrame):
            df = value.copy()
            df.columns = df.columns.str.strip()

            columnas_requeridas = ["Tierra Superficie: : m²", "Paredes sin apertura: m²"]
            if not all(col in df.columns for col in columnas_requeridas):
                continue

            for _, row in df.iterrows():
                try:
                    nombre_habitacion = row.iloc[0]  # Primera columna es el nombre

                    superficie = float(row.get("Tierra Superficie: : m²", 0) or 0)
                    paredes_sin_apertura = float(row.get("Paredes sin apertura: m²", 0) or 0)
                    perimetro_interno = float(row.get("Tierra Perímetro: m", 0) or 0)
                    perimetro_techo = float(row.get("Techo Perímetro: m", 0) or 0)
                    diferencia = abs(perimetro_interno - perimetro_techo)
                    techo = superficie * 1.15 if diferencia >= 0.1 else superficie

                    resultados[nombre_habitacion] = {
                        "MAGICPLAN - ÁREA PISO": superficie,
                        "MAGICPLAN - ÁREA PARED": paredes_sin_apertura,
                        "MAGICPLAN - ÁREA CUBIERTA": techo,
                        "MAGICPLAN - PERIMETRO PISO": perimetro_interno,
                        "MAGICPLAN - PERIMETRO CUBIERTA": perimetro_techo,
                    }
                    
                except Exception as e:
                    resultados[f"Error en {tabla_key}"] = f"Error al procesar habitación: {str(e)}"

    return resultados

@st.cache_data
def load_pdf(file):
    return file.read()

@st.cache_data
def load_image(file):
    return Image.open(file)

def inicio():

    st.title("Ingreso de archivos")
    st.write("Cargue los archivos correspondientes a la vivienda.")

    # Carga automática del archivo Excel sin necesidad de subirlo manualmente
    try:
        st.session_state["costos_excel"] = load_excel_local()
        st.success("Archivo Excel de costos cargado correctamente desde el código.")
    except Exception as e:
        st.error(f"Error al cargar el archivo Excel: {str(e)}")
        
        
    try:
        st.session_state["beneficiarios_excel"] = load_beneficiarios()
        st.success("Archivo de beneficiarios cargado correctamente.")
    except Exception as e:
        st.error(f"Error al cargar beneficiarios: {e}")
    

    # Cargar archivos desde la interfaz web
    plano_file = st.file_uploader("Sube un archivo (Plano o Imagen)", type=["pdf", "png", "jpg", "jpeg"])
    resultados_csv = st.file_uploader("Sube un archivo CSV (Resultados MagicPlan)", type=["csv"])

    # Validar que ambos archivos sean subidos antes de continuar
    if resultados_csv:
        # Procesar el CSV
        tablas, codigo = procesar_csv_bytes(resultados_csv)
        st.session_state["resultados_csv"] = calcular_propiedades_habitacion(tablas)
        st.success("Archivo CSV cargado correctamente.")

        # Si existe el archivo de plano (PDF/imagen), lo procesamos. Si no, lo ignoramos.
        if plano_file:
            file_extension = plano_file.name.split(".")[-1].lower()

            if file_extension == "pdf":
                st.session_state["plano_pdf"] = load_pdf(plano_file)
                st.success("Archivo PDF cargado correctamente.")
            elif file_extension in ["png", "jpg", "jpeg"]:
                st.session_state["plano_img"] = load_image(plano_file)
                st.success("Imagen cargada correctamente.")
        else:
            st.info("No se ha subido archivo de plano; se continuará sin visualizar un plano.")

    else:
        # Aviso de que el CSV es obligatorio
        st.warning("⚠️ Debe subir el archivo CSV para continuar.")

@st.cache_data
def load_image(file):
    return Image.open(file)

@st.cache_data
def load_beneficiarios():
    ruta = "Base_Beneficiarios.xlsx"  # Ajusta si lo tienes en otra carpeta
    df = pd.read_excel(ruta, sheet_name="Hoja1")
    return df

def obtener_datos_beneficiario(cedula, df_benef):
    df_filtrado = df_benef[df_benef["C.C:"] == int(cedula)]
    if df_filtrado.empty:
        return None
    fila = df_filtrado.iloc[0]
    return {
        "cedula": cedula,
        "nombre": fila["NOMBRE BENEFICIARIO:"],
        "direccion": fila["DIRECCION:"],
        "telefono": fila["TELEFONO:"],
        "idhogar": fila["ID_HOGAR:"]
    }



def main():
    
    st.set_page_config(page_title="Modificación de vivienda", layout="wide")
    
    st.sidebar.markdown("<h2 style='text-align: center; color: green;'>💰 Costo total permitido: $13.201.188</h2>", unsafe_allow_html=True)

    st.sidebar.markdown("### Buscar Beneficiario")

    cedula_input = st.sidebar.text_input("Ingrese la cédula:", key="input_cedula")
    buscar_btn = st.sidebar.button("Buscar")

    if buscar_btn:
        if not cedula_input.strip():
            st.sidebar.warning("Ingrese una cédula válida.")
        else:
            st.session_state["cedula_usuario"] = cedula_input.strip()

            # Buscar en el Excel de beneficiarios
            if "beneficiarios_excel" in st.session_state:
                datos_benef = obtener_datos_beneficiario(
                    st.session_state["cedula_usuario"],
                    st.session_state["beneficiarios_excel"]
                )
                if datos_benef:
                    st.session_state["datos_beneficiario"] = datos_benef
                    st.sidebar.success("Beneficiario encontrado.")
                    st.sidebar.markdown(f"**Nombre:** {datos_benef['nombre']}")
                    st.sidebar.markdown(f"**Dirección:** {datos_benef['direccion']}")
                    st.sidebar.markdown(f"**Teléfono:** {datos_benef['telefono']}")

                else:
                    st.sidebar.error("No se encontró la cédula en la base.")

    
    if st.sidebar.button("Reiniciar aplicación"):
        st.session_state.clear()  # Limpia todos los valores almacenados
        st.rerun()
    
    # 🔹 Valor máximo permitido fijo
    max_total = 13201188  # 15.600.000
    
    # Inicializar 'max_costo' en st.session_state si no existe
    if "max_costo" not in st.session_state:
        st.session_state["max_costo"] = max_total
    
    
    # En la parte de la interfaz (por ejemplo, en la función main() o en la sección de la barra lateral)
    selected = st.sidebar.selectbox(
        "Seleccione un Técnico",
        options=[tec["PROFESIONAL"] for tec in TECNICOS]
    )
    if selected:
        # Buscamos el diccionario del técnico seleccionado
        tecnico_seleccionado = next((tec for tec in TECNICOS if tec["PROFESIONAL"] == selected), None)
        st.session_state["selected_tecnico"] = tecnico_seleccionado
        st.sidebar.success(f"Técnico seleccionado: {tecnico_seleccionado['PROFESIONAL']}")


    

    # 🔹 Continuar con las pantallas de la aplicación
    inicio()
    vista_archivos(st.session_state['max_costo'])

@st.cache_data
def load_pdf(file):
    return file.read()

@st.cache_data
def load_csv(file):
    return pd.read_csv(file)

# Ruta del archivo Excel local (ajusta esto a tu ubicación real)
RUTA_ARCHIVO_COSTOS = "TURBO_ARCHIVO_PARA_TRABAJAR.xlsx"

# Función para cargar el archivo Excel desde la ruta local
@st.cache_data
def load_excel_local():
    return pd.read_excel(RUTA_ARCHIVO_COSTOS, sheet_name="FORMATO DE OFERTA ECONÓMICA")


def ultimas_dos_palabras(texto: str) -> str:
    palabras = texto.split()  # Dividir el texto en palabras
    return " ".join(palabras[-2:]) if len(palabras) >= 2 else texto

def verificar_palabras(texto, lista_referencia):
    palabras = {palabra.strip() for palabra in texto.split(",")}  # Convertir en conjunto sin espacios extra
    lista_referencia = set(lista_referencia)  # Convertir la lista en conjunto
    return not palabras.isdisjoint(lista_referencia)

def vista_archivos(max_total):
    st.title("Modificaciones a realizar")

    # Mostrar el archivo PDF si se ha subido
    if "plano_pdf" in st.session_state:
        st.subheader("Plano PDF")
        pdf_viewer(st.session_state["plano_pdf"], width="50%")
    # Mostrar la imagen si el usuario subió una imagen en lugar de un PDF
    elif "plano_img" in st.session_state:
        st.subheader("Plano en Imagen")
        st.image(st.session_state["plano_img"], caption="Plano en imagen", use_container_width=True)

    # Si los archivos CSV y Excel están cargados, mostrar la interfaz de modificaciones
    if "resultados_csv" in st.session_state and "costos_excel" in st.session_state:
        st.subheader("Selección de Habitaciones")
        cedula_filtro = st.session_state.get("cedula_usuario", "").strip()

        habitaciones = [
            key for key in st.session_state["resultados_csv"].keys()
            if key.strip().startswith(cedula_filtro + " ") and len(key.strip().split()) > 1 and "piso" not in key.lower()
        ]

        if not habitaciones:
            st.warning("No hay habitaciones asociadas a la cédula ingresada.")
            return

        actividades = st.session_state["costos_excel"]
        estados = {}
        subtotales = {}

        for habitacion in habitaciones:
            activo = habitacion.startswith("#")
            estados[habitacion] = st.checkbox(habitacion, value=activo, key=f"habitacion_{habitacion}")
            subtotal = 0.0

            if estados[habitacion]:
                st.subheader(f"🏠 Modificaciones de {habitacion}")  # Sin expander para cada habitación

                # Crear diccionario para almacenar categorías con actividades
                categorias_actividades = {}
                categoria_actual = None

                for _, row in actividades.iterrows():
                    actividad = row.get("ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS", "")
                    unidad = row.get("Unidad", None)
                    item = row.get("Item", "")
                    valor_unitario = row.get("Valor Unitario ofertado (**)", 0.0)
                    medicion = row.get("ÁREA", "")
                    formula = row.get("FORMULA", "")
                    formula = "" if pd.isna(formula) else formula

                    # Si es una categoría (título en mayúsculas)
                    if actividad.isupper():
                        categoria_actual = actividad
                        categorias_actividades[categoria_actual] = []
                    elif categoria_actual:
                        categorias_actividades[categoria_actual].append(
                            (item, actividad, unidad, valor_unitario, medicion, formula)
                        )

                # Mostrar las categorías con sus actividades en expander
                for categoria, lista_actividades in categorias_actividades.items():
                    with st.expander(f"📂 {categoria}", expanded=False):
                        for item, actividad, unidad, valor_unitario, medicion, formula in lista_actividades:
                            check = st.checkbox(
                                f"{item} -- {actividad} [Unidad: {unidad}] (Precio unitario: ${valor_unitario:,.2f})",
                                key=f"check_{habitacion}_{actividad}"
                            )

                            if check:
                                cantidad_key = f"cantidad_{habitacion}_{actividad}"
                                valor_guardado_key = f"valor_{habitacion}_{actividad}"
                                if valor_guardado_key not in st.session_state:
                                    st.session_state[valor_guardado_key] = 0.0

                                # Si es una actividad manual (ingreso del usuario)
                                if "USUARIO" in medicion.upper():
                                    # Usar text_input para iniciar vacío
                                    cantidad_str = st.text_input(
                                        f"Ingrese la cantidad ({unidad}).",
                                        value="",  # Inicia vacío
                                        key=cantidad_key,
                                        placeholder="Ej: 2.50"
                                    )
                                    if st.button(f"Guardar cantidad", key=f"button_{habitacion}_{actividad}"):
                                        if cantidad_str.strip():
                                            try:
                                                cantidad = float(cantidad_str)
                                                st.session_state[valor_guardado_key] = cantidad * valor_unitario
                                                st.success(
                                                    f"Valor guardado para {actividad}: "
                                                    f"${st.session_state[valor_guardado_key]:,.2f}"
                                                )
                                            except ValueError:
                                                st.error("Por favor, ingresa un número válido.")
                                        else:
                                            st.warning("No ingresaste ningún valor.")

                                else:
                                    # Resto de las ramas: actividades que vienen de MagicPlan
                                    if "ALTURA" in formula:    
                                        cantidad = st.number_input(
                                            f"Valor MagicPlan ({ultimas_dos_palabras(medicion)})",
                                            value=st.session_state["resultados_csv"][habitacion][medicion],
                                            min_value=0.0,
                                            key=cantidad_key
                                        )
                                        valor_input_str = st.text_input(
                                            "Ingrese la altura (metros).",
                                            value="",
                                            key=f"{cantidad_key}_aux_txt",
                                            placeholder="Ej: 2.40"
                                        )
                                        if st.button(f"Guardar cantidad", key=f"button_{habitacion}_{actividad}"):
                                            if not valor_input_str.strip():
                                                st.warning("No has ingresado ningún valor para la altura.")
                                            else:
                                                try:
                                                    valor_input_float = float(valor_input_str)
                                                    st.session_state[valor_guardado_key] = (
                                                        cantidad * valor_unitario * valor_input_float
                                                    )
                                                    st.success(
                                                        f"Valor guardado para {actividad}: "
                                                        f"${st.session_state[valor_guardado_key]:,.2f}"
                                                    )
                                                except ValueError:
                                                    st.error("Por favor, ingresa un número válido para la altura.")
                                    elif formula != "":
                                        cantidad = st.number_input(
                                            f"Ingrese la cantidad ({unidad}).",
                                            value=st.session_state["resultados_csv"][habitacion][medicion],
                                            min_value=0.0,
                                            key=cantidad_key
                                        )
                                        if st.button(f"Guardar cantidad", key=f"button_{habitacion}_{actividad}"):
                                            st.session_state[valor_guardado_key] = cantidad * valor_unitario
                                            st.success(
                                                f"Valor guardado para {actividad}: "
                                                f"${st.session_state[valor_guardado_key]:,.2f}"
                                            )
                                    else:    
                                        cantidad = st.number_input(
                                            f"Valor MagicPlan ({ultimas_dos_palabras(medicion)}) [Unidad: {unidad}]",
                                            value=st.session_state["resultados_csv"][habitacion][medicion],
                                            min_value=0.0,
                                            key=cantidad_key
                                        )
                                        st.session_state[valor_guardado_key] = cantidad * valor_unitario
                                        st.success(
                                            f"Valor guardado para {actividad}: "
                                            f"${st.session_state[valor_guardado_key]:,.2f}"
                                        )

                                subtotal += st.session_state[valor_guardado_key]

            subtotales[habitacion] = subtotal

        # ─────────────────────────────────────────────────────────────────────────────
        # FORMATEAR SUBTOTALES COMO MONEDA (SIN DECIMALES) EN LA TABLA
        # ─────────────────────────────────────────────────────────────────────────────
        total_general = sum(subtotales.values())
        st.sidebar.subheader("Subtotales por Habitación")

        # Convertir 'subtotales' en DataFrame
        df_subtotales = pd.DataFrame(list(subtotales.items()), columns=["Habitación", "Subtotal ($)"])
        
        # Redondear y formatear
        df_subtotales["Subtotal ($)"] = df_subtotales["Subtotal ($)"].round(0).astype(int)
        df_subtotales["Subtotal ($)"] = df_subtotales["Subtotal ($)"].apply(lambda x: f"${x:,.0f}")

        st.sidebar.dataframe(df_subtotales, hide_index=True)

        st.sidebar.subheader("Total General")
        if total_general > max_total:
            st.sidebar.markdown(
                f"<span style='color: red; font-weight: bold;'>Total: ${total_general:,.0f}</span>",
                unsafe_allow_html=True
            )
            st.sidebar.warning('Se ha superado el monto máximo permisible.')
        else:
            st.sidebar.markdown(f"Total: ${total_general:,.0f}")
            obtener_tabla_habitaciones()
            if "export_excel" in st.session_state and total_general > 0:
                try:
                    with open(st.session_state["export_excel"], "rb") as file:
                        st.sidebar.download_button(
                            label="Descargar Reporte",
                            data=file,
                            file_name="Reporte_Resultado.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.sidebar.error(f"Error al generar el archivo: {str(e)}")
    else:
        st.warning('Ingrese los archivos para iniciar el proceso, en la sección Inicio.')



        
def registro_login():
    st.title("Registro o Inicio de Sesión")
    opcion = st.radio("Elige una opción:", ["Iniciar Sesión", "Registrarse"])
    
    if opcion == "Iniciar Sesión":
        usuario = st.text_input("Usuario")
        contraseña = st.text_input("Contraseña", type="password")
        if st.button("Ingresar"):
            st.success(f"Bienvenido, {usuario}!")
    
    elif opcion == "Registrarse":
        nuevo_usuario = st.text_input("Nuevo Usuario")
        nueva_contraseña = st.text_input("Nueva Contraseña", type="password")
        confirmar_contraseña = st.text_input("Confirmar Contraseña", type="password")
        if st.button("Registrarse"):
            if nueva_contraseña == confirmar_contraseña:
                st.success("Registro exitoso. Ahora puedes iniciar sesión.")
            else:
                st.error("Las contraseñas no coinciden.")

if __name__ == "__main__":
    main()